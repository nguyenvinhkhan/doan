from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from datetime import datetime, date, timezone, timedelta

VN_TZ = timezone(timedelta(hours=7))  # UTC+7

def _to_vn(dt):
    """Chuyển datetime từ UTC sang giờ Việt Nam."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(VN_TZ)
from io import BytesIO
from database import get_db
from auth import get_current_user
import models
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter()


def _border(style="thin"):
    s = Side(style=style, color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_fill():
    return PatternFill("solid", start_color="0D1B2A", end_color="0D1B2A")

def _subheader_fill():
    return PatternFill("solid", start_color="1A2E42", end_color="1A2E42")

def _alt_fill():
    return PatternFill("solid", start_color="F2F7FF", end_color="F2F7FF")

def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def _apply_header(cell, text, font_size=11, bold=True, color="00E5FF", fill=None):
    cell.value = text
    cell.font = Font(name="Arial", bold=bold, size=font_size, color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()
    if fill:
        cell.fill = fill
    else:
        cell.fill = _header_fill()

def _apply_cell(cell, value, bold=False, color="000000", align="left", fill=None, num_format=None):
    cell.value = value
    cell.font = Font(name="Arial", bold=bold, size=10, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _border()
    if fill:
        cell.fill = fill
    if num_format:
        cell.number_format = num_format


@router.get("/attendance")
def export_attendance(
    month: Optional[int] = Query(None),
    year:  Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Xuất báo cáo điểm danh tháng ra Excel."""
    today = date.today()
    m = month or today.month
    y = year or today.year
    prefix = f"{y}-{m:02d}"

    records = (
        db.query(models.Attendance)
        .filter(models.Attendance.date.like(f"{prefix}%"))
        .order_by(models.Attendance.date, models.Attendance.check_in)
        .all()
    )

    employees = {e.id: e for e in db.query(models.Employee).all()}

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════
    # Sheet 1: Chi tiết điểm danh
    # ══════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Chi Tiết Điểm Danh"
    ws1.sheet_view.showGridLines = False
    ws1.row_dimensions[1].height = 40

    # Tiêu đề
    ws1.merge_cells("A1:I1")
    title_cell = ws1["A1"]
    title_cell.value = f"BÁO CÁO ĐIỂM DANH THÁNG {m:02d}/{y}"
    title_cell.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", start_color="00335C", end_color="00335C")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws1.row_dimensions[2].height = 14  # spacer
    ws1.row_dimensions[3].height = 30

    headers = ["#", "Mã NV", "Họ và Tên", "Phòng Ban", "Ngày", "Giờ Vào", "Giờ Ra", "Trạng Thái", "Độ Chính Xác"]
    for col, h in enumerate(headers, 1):
        _apply_header(ws1.cell(row=3, column=col), h)

    status_colors = {"present": "00AA44", "late": "CC8800", "absent": "CC3300"}
    status_labels = {"present": "Đúng Giờ", "late": "Đi Trễ", "absent": "Vắng"}

    for i, rec in enumerate(records, 1):
        row = i + 3
        emp = employees.get(rec.employee_id)
        fill = _alt_fill() if i % 2 == 0 else None

        _apply_cell(ws1.cell(row=row, column=1), i, align="center", fill=fill)
        _apply_cell(ws1.cell(row=row, column=2), emp.employee_code if emp else "-", align="center", fill=fill)
        _apply_cell(ws1.cell(row=row, column=3), emp.full_name if emp else "-", fill=fill)
        _apply_cell(ws1.cell(row=row, column=4), emp.department if emp else "-", fill=fill)
        _apply_cell(ws1.cell(row=row, column=5), rec.date, align="center", fill=fill)
        _apply_cell(ws1.cell(row=row, column=6),
                    _to_vn(rec.check_in).strftime("%H:%M:%S") if rec.check_in else "-",
                    align="center", fill=fill)
        _apply_cell(ws1.cell(row=row, column=7),
                    _to_vn(rec.check_out).strftime("%H:%M:%S") if rec.check_out else "-",
                    align="center", fill=fill)
        status_col = ws1.cell(row=row, column=8)
        label = status_labels.get(rec.status, rec.status)
        color = status_colors.get(rec.status, "333333")
        _apply_cell(status_col, label, bold=True, color=color, align="center", fill=fill)
        _apply_cell(ws1.cell(row=row, column=9),
                    f"{rec.confidence*100:.1f}%" if rec.confidence else "-",
                    align="center", fill=fill)

    _set_col_widths(ws1, {"A":"6","B":"12","C":"22","D":"18","E":"12","F":"12","G":"12","H":"14","I":"14"})

    # ══════════════════════════════════════════════
    # Sheet 2: Thống kê nhân viên
    # ══════════════════════════════════════════════
    ws2 = wb.create_sheet("Thống Kê Nhân Viên")
    ws2.sheet_view.showGridLines = False
    ws2.row_dimensions[1].height = 40
    ws2.merge_cells("A1:G1")
    c = ws2["A1"]
    c.value = f"THỐNG KÊ NHÂN VIÊN THÁNG {m:02d}/{y}"
    c.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="00335C", end_color="00335C")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws2.row_dimensions[3].height = 30
    for col, h in enumerate(["#","Mã NV","Họ và Tên","Phòng Ban","Đúng Giờ","Đi Trễ","Vắng"], 1):
        _apply_header(ws2.cell(row=3, column=col), h)

    emp_stats = {}
    for rec in records:
        eid = rec.employee_id
        if eid not in emp_stats:
            emp_stats[eid] = {"present": 0, "late": 0, "absent": 0}
        emp_stats[eid][rec.status] = emp_stats[eid].get(rec.status, 0) + 1

    for i, (eid, stats) in enumerate(emp_stats.items(), 1):
        row = i + 3
        emp = employees.get(eid)
        fill = _alt_fill() if i % 2 == 0 else None
        _apply_cell(ws2.cell(row=row, column=1), i, align="center", fill=fill)
        _apply_cell(ws2.cell(row=row, column=2), emp.employee_code if emp else "-", align="center", fill=fill)
        _apply_cell(ws2.cell(row=row, column=3), emp.full_name if emp else "-", fill=fill)
        _apply_cell(ws2.cell(row=row, column=4), emp.department if emp else "-", fill=fill)
        _apply_cell(ws2.cell(row=row, column=5), stats["present"], align="center", color="00AA44", bold=True, fill=fill)
        _apply_cell(ws2.cell(row=row, column=6), stats["late"],    align="center", color="CC8800", bold=True, fill=fill)
        _apply_cell(ws2.cell(row=row, column=7), stats["absent"],  align="center", color="CC3300", bold=True, fill=fill)

    # Tổng
    total_row = len(emp_stats) + 4
    ws2.cell(row=total_row, column=3).value = "TỔNG"
    ws2.cell(row=total_row, column=3).font = Font(name="Arial", bold=True, size=10)
    _total_colors = {5: "00AA44", 6: "CC8800", 7: "CC3300"}
    for col, key in [(5,"present"),(6,"late"),(7,"absent")]:
        val = sum(s[key] for s in emp_stats.values())
        c = ws2.cell(row=total_row, column=col)
        c.value = val
        c.font = Font(name="Arial", bold=True, size=10, color=_total_colors[col])
        c.alignment = Alignment(horizontal="center")
        c.border = _border()
        c.fill = PatternFill("solid", start_color="E8F0FF", end_color="E8F0FF")

    _set_col_widths(ws2, {"A":"6","B":"12","C":"22","D":"18","E":"12","F":"12","G":"12"})

    # ══════════════════════════════════════════════
    # Sheet 3: Tổng hợp theo ngày
    # ══════════════════════════════════════════════
    ws3 = wb.create_sheet("Tổng Hợp Theo Ngày")
    ws3.sheet_view.showGridLines = False
    ws3.row_dimensions[1].height = 40
    ws3.merge_cells("A1:E1")
    c = ws3["A1"]
    c.value = f"TỔNG HỢP THEO NGÀY - THÁNG {m:02d}/{y}"
    c.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="00335C", end_color="00335C")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws3.row_dimensions[3].height = 30
    for col, h in enumerate(["Ngày","Đúng Giờ","Đi Trễ","Vắng","Tổng"], 1):
        _apply_header(ws3.cell(row=3, column=col), h)

    daily = {}
    for rec in records:
        d = rec.date
        if d not in daily:
            daily[d] = {"present":0,"late":0,"absent":0}
        daily[d][rec.status] = daily[d].get(rec.status,0) + 1

    for i, (d, stats) in enumerate(sorted(daily.items()), 1):
        row = i + 3
        fill = _alt_fill() if i % 2 == 0 else None
        _apply_cell(ws3.cell(row=row,column=1), d, align="center", fill=fill)
        _apply_cell(ws3.cell(row=row,column=2), stats["present"], align="center", color="00AA44", bold=True, fill=fill)
        _apply_cell(ws3.cell(row=row,column=3), stats["late"],    align="center", color="CC8800", bold=True, fill=fill)
        _apply_cell(ws3.cell(row=row,column=4), stats["absent"],  align="center", color="CC3300", bold=True, fill=fill)
        total = stats["present"]+stats["late"]+stats["absent"]
        _apply_cell(ws3.cell(row=row,column=5), total, align="center", bold=True, fill=fill)

    _set_col_widths(ws3, {"A":"14","B":"12","C":"12","D":"12","E":"12"})

    # Xuất file
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"diem_danh_{y}_{m:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )